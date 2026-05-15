"""
Gera resultados_adicionais.xlsx com os resultados do VRPSolver para as 20
instâncias adicionais (variant_1 + variant_2).

Abas:
  - "variant_1"        : resultados das 10 instâncias da pasta variant_1
  - "variant_2"        : resultados das 10 instâncias da pasta variant_2
  - "Comparativo v1xv2": instância-base lado a lado, com Δ% de custo, tempo e #colunas
  - "Notas"            : metodologia / legendas

Colunas (por instância) em cada aba de variante:
  Instância | Best solution | Integrality gap (%) | Solution time (s)
  | Iterations (CG) | Nodes | Columns generated (total) | Columns pool (peak)
  | Root LB | Optimal? | Wall time (s)

A "Comparativo v1xv2" pareia variant_1/instance_k com variant_2/variant2_instance_k.
"""

from __future__ import annotations

from pathlib import Path
import re
import csv

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

ROOT = Path(r"C:\Users\porin\OneDrive\Documentos\Python-Mestrado\Modelagem Matemática\Programação Inteira - Uchoa\Problema VRPRDL\VRPRDL_VRPSolver_")
OUT_XLSX = ROOT / "resultados_adicionais.xlsx"

VARIANTS = {
    "variant_1": {
        "log_dir": ROOT / "logs_extra_v1",
        "summary": ROOT / "logs_extra_v1" / "_summary.csv",
        "rx_id":   re.compile(r"^instance_(\d+)$"),
    },
    "variant_2": {
        "log_dir": ROOT / "logs_extra_v2",
        "summary": ROOT / "logs_extra_v2" / "_summary.csv",
        "rx_id":   re.compile(r"^variant2_instance_(\d+)$"),
    },
}


def parse_log(path: Path) -> dict:
    cg_iters = 0
    total_cols = 0
    pool_peak = 0
    rx_cg = re.compile(r"<DWph=\s*\d+>\s*<it=\s*\d+>.*?<nCl=\s*(\d+)>")
    rx_pool = re.compile(r"(\d+)\s+columns\s+\(")
    if not path.exists():
        return {"iterations": None, "columns_generated": None, "columns_pool_peak": None}
    with path.open("r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            m_cg = rx_cg.search(line)
            if m_cg:
                cg_iters += 1
                total_cols += int(m_cg.group(1))
            m_pool = rx_pool.search(line)
            if m_pool:
                n = int(m_pool.group(1))
                if n > pool_peak:
                    pool_peak = n
    return {
        "iterations": cg_iters,
        "columns_generated": total_cols,
        "columns_pool_peak": pool_peak,
    }


def _to_float(s):
    try:
        return float(s)
    except Exception:
        return None


def load_variant_rows(variant: str) -> list[dict]:
    cfg = VARIANTS[variant]
    rows: list[dict] = []
    if not cfg["summary"].exists():
        print(f"[{variant}] arquivo não encontrado: {cfg['summary']}")
        return rows
    with cfg["summary"].open("r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for r in reader:
            name = r["instance"]
            m = cfg["rx_id"].match(name)
            if not m:
                continue
            k = int(m.group(1))
            inc = _to_float(r.get("best_inc"))
            root = _to_float(r.get("root_db"))
            gap = None
            if inc is not None and root is not None and root > 0:
                gap = round((inc - root) / root * 100.0, 4)
            log_path = cfg["log_dir"] / f"{name}.log"
            log_stats = parse_log(log_path)
            rows.append(
                {
                    "_id": k,
                    "_name": name,
                    "Instância": name,
                    "Best solution": int(inc) if inc is not None else None,
                    "Integrality gap (%)": gap,
                    "Solution time (s)": round(_to_float(r.get("total_time_s")) or 0.0, 2),
                    "Iterations (CG)": log_stats["iterations"],
                    "Nodes": int(r.get("nb_nodes", 0) or 0),
                    "Columns generated (total)": log_stats["columns_generated"],
                    "Columns pool (peak)":      log_stats["columns_pool_peak"],
                    "Root LB": round(root, 2) if root is not None else None,
                    "Optimal?": "yes" if r.get("optimal") == "1" else "no",
                    "Wall time (s)": int(_to_float(r.get("wall_time_s")) or 0),
                }
            )
    rows.sort(key=lambda x: x["_id"])
    return rows


def build_compare(rows_v1: list[dict], rows_v2: list[dict]) -> pd.DataFrame:
    by_id_v1 = {r["_id"]: r for r in rows_v1}
    by_id_v2 = {r["_id"]: r for r in rows_v2}
    ids = sorted(set(by_id_v1) | set(by_id_v2))

    records = []
    for k in ids:
        a = by_id_v1.get(k)
        b = by_id_v2.get(k)
        cost_a = a["Best solution"] if a else None
        cost_b = b["Best solution"] if b else None
        time_a = a["Solution time (s)"] if a else None
        time_b = b["Solution time (s)"] if b else None
        cols_a = a["Columns generated (total)"] if a else None
        cols_b = b["Columns generated (total)"] if b else None

        def pct(x, y):
            if x is None or y is None or y == 0:
                return None
            return round((x - y) / y * 100.0, 2)

        records.append(
            {
                "ID base": k,
                "v1 instância": a["_name"] if a else None,
                "v2 instância": b["_name"] if b else None,
                "Custo v1": cost_a,
                "Custo v2": cost_b,
                "Δ Custo (v2 vs v1) (%)": pct(cost_b, cost_a),
                "Tempo v1 (s)": time_a,
                "Tempo v2 (s)": time_b,
                "Δ Tempo (v2 vs v1) (%)": pct(time_b, time_a),
                "Colunas v1": cols_a,
                "Colunas v2": cols_b,
                "Δ Colunas (v2 vs v1) (%)": pct(cols_b, cols_a),
                "Optimal v1": a["Optimal?"] if a else None,
                "Optimal v2": b["Optimal?"] if b else None,
            }
        )
    return pd.DataFrame(records)


def df_for_sheet(rows: list[dict]) -> pd.DataFrame:
    df = pd.DataFrame([{k: v for k, v in r.items() if not k.startswith("_")} for r in rows])
    return df


def write_xlsx(rows_v1: list[dict], rows_v2: list[dict], df_cmp: pd.DataFrame):
    df_v1 = df_for_sheet(rows_v1)
    df_v2 = df_for_sheet(rows_v2)

    notas = pd.DataFrame(
        {
            "Item": [
                "Conjunto de instâncias",
                "Origem dos dados",
                "Best solution",
                "Integrality gap (%)",
                "Solution time (s)",
                "Iterations (CG)",
                "Nodes",
                "Columns generated (total)",
                "Columns pool (peak)",
                "Root LB",
                "Optimal?",
                "Wall time (s)",
                "Comparativo v1xv2",
                "Time limit do batch",
                "Hardware",
                "Formulação",
            ],
            "Valor / Observação": [
                "20 instâncias adicionais (10 em variant_1 + 10 em variant_2), n = 40 clientes em todas",
                "instancias_turco/additional_instances/{variant_1, variant_2}/*.txt",
                "Custo incumbente final = :bcRecBestInc do VrpSolver",
                "(best_inc - root_db)/root_db * 100; root_db = :bcRecRootDb",
                ":bcTimeMain do VrpSolver (tempo total de solve)",
                "Contagem de linhas '<DWph=' no log do VrpSolver (iterações de CG)",
                ":bcCountNodeProc (nós processados no B&P)",
                "Soma de <nCl=N> em todas as iterações de CG (total acumulado de colunas geradas)",
                "Máximo observado de 'N columns (...)' no log (pico do tamanho do pool)",
                "Bound da raiz do B&P (limite inferior dual após resolver a relaxação)",
                "yes/no — todas as 20 chegaram ao ótimo",
                "Tempo de relógio (wall) por instância, incluindo startup do Julia",
                "Pareia variant_1/instance_k com variant_2/variant2_instance_k pelo mesmo k base.",
                "TIMEOUT_S=7200s por instância (nenhuma chegou perto do limite)",
                "Execução em WSL (Ubuntu sobre Windows) na máquina do usuário",
                "VrpSolver (Pessoa/Sadykov/Uchoa) com packing sets por cliente; recursos: capacidade e tempo",
            ],
        }
    )

    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        df_v1.to_excel(writer,  sheet_name="variant_1",         index=False)
        df_v2.to_excel(writer,  sheet_name="variant_2",         index=False)
        df_cmp.to_excel(writer, sheet_name="Comparativo v1xv2", index=False)
        notas.to_excel(writer,  sheet_name="Notas",             index=False)

    wb = load_workbook(OUT_XLSX)
    header_fill = PatternFill("solid", fgColor="DCE6F1")
    header_font = Font(bold=True)

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        max_col = ws.max_column
        for col_idx in range(1, max_col + 1):
            letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[letter]:
                val = "" if cell.value is None else str(cell.value)
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 40)
        ws.freeze_panes = "A2"

    wb.save(OUT_XLSX)


def main():
    rows_v1 = load_variant_rows("variant_1")
    rows_v2 = load_variant_rows("variant_2")
    df_cmp = build_compare(rows_v1, rows_v2)
    print(f"variant_1: {len(rows_v1)} linhas;  variant_2: {len(rows_v2)} linhas")
    write_xlsx(rows_v1, rows_v2, df_cmp)
    print(f"Arquivo gerado: {OUT_XLSX}")


if __name__ == "__main__":
    main()
