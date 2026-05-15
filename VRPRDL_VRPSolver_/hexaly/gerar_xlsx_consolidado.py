# -*- coding: utf-8 -*-
"""Gera resultados_consolidado_3solvers.xlsx:

- Aba 1 "Originais (Tab.3)" : 40 instancias originais comparando
  Tabela 3 do Ozbaygin (2017) x VrpSolver (nosso) x Hexaly (nosso).
- Aba 2 "Variantes (Tab.7)" : 20 instancias adicionais (10 v1 + 10 v2)
  comparando Tabela 7 do Ozbaygin (2017) x VrpSolver x Hexaly.
- Aba 3 "Resumo"            : medias / maximos / counts por familia.
- Aba 4 "Notas"             : metodologia, mapeamento de instancias e legendas.

Mapeamentos:
  instance_{k-1}-triangle           (nosso) <-> linha k da Tabela 3 (k=1..40).
  variant_1/instance_K              (nosso) <-> "{41+K}_v1" da Tabela 7 (K=0..9).
  variant_2/variant2_instance_K     (nosso) <-> "{41+K}_v2" da Tabela 7 (K=0..9).
"""

from __future__ import annotations
from pathlib import Path
import csv
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
HEX_DIR = ROOT / "hexaly"
HEX_SUMMARY = HEX_DIR / "logs_hexaly" / "_summary.csv"

VRPS_ORIG = ROOT / "logs" / "_summary.csv"
VRPS_V1 = ROOT / "logs_extra_v1" / "_summary.csv"
VRPS_V2 = ROOT / "logs_extra_v2" / "_summary.csv"

OUT_XLSX = ROOT / "resultados_consolidado_3solvers.xlsx"


OZ_TL_S_BY_K = {**{k: 7200 for k in range(1, 31)}, **{k: 21600 for k in range(31, 41)}}
# (k, init, best, gap_pct, time_s_or_None=TL, iters, nodes)
OZBAYGIN_TABLE3 = [
    (1,  2074, 901,   0.0,   0.58,   39,    1),
    (2,  2316, 1286,  0.0,   0.27,   29,    1),
    (3,  2234, 991,   0.0,   0.43,   30,    1),
    (4,  1982, 1062,  0.0,   0.27,   30,    1),
    (5,  3322, 1832,  0.0,   0.04,   26,    1),
    (6,  3328, 1294,  1.73,  2.30,   245,   31),
    (7,  3204, 1155,  1.29,  30.47,  773,   70),
    (8,  3170, 1455,  0.0,   0.18,   39,    1),
    (9,  2838, 1260,  1.86,  3.32,   307,   32),
    (10, 3270, 1684,  0.0,   0.55,   33,    1),
    (11, 4932, 1922,  0.31,  14.28,  283,   27),
    (12, 4610, 2324,  2.52,  120.38, 16914, 2707),
    (13, 4868, 1747,  0.0,   23.33,  101,   1),
    (14, 4084, 1273,  0.0,   30.83,  100,   1),
    (15, 4656, 1694,  0.0,   22.78,  97,    1),
    (16, 4770, 1938,  0.0,   57.13,  82,    1),
    (17, 4502, 1965,  0.10,  5.75,   117,   3),
    (18, 5392, 1827,  0.0,   2.10,   83,    1),
    (19, 5286, 2083,  2.46,  93.55,  4140,  413),
    (20, 4236, 1822,  0.0,   78.70,  103,   1),
    (21, 10374, 3761, 0.0,   606.60, 216,   1),
    (22, 9316, 2828,  0.0,   272.43, 357,   3),
    (23, 10326, 4440, 0.01,  238.27, 208,   3),
    (24, 10536, 3378, 0.0,   382.37, 250,   1),
    (25, 9784, 9784,  None,  None,   1,     0),
    (26, 10822, 4536, 0.0,   78.61,  158,   1),
    (27, 10634, 2865, 0.0,   296.23, 242,   1),
    (28, 9450, 9450,  None,  None,   1,     0),
    (29, 10988, 3964, 0.73,  None,   82536, 11593),
    (30, 11840, 4107, 0.0,   45.22,  177,   1),
    (31, 18142, 18142, None, None,   1,     0),
    (32, 19514, 19514, None, None,   1,     0),
    (33, 18008, 18008, None, None,   4,     0),
    (34, 19880, 19880, None, None,   2,     0),
    (35, 19196, 19196, None, None,   1,     0),
    (36, 21772, 21772, None, None,   3,     0),
    (37, 20010, 20010, None, None,   2,     0),
    (38, 20032, 20032, None, None,   480,   0),
    (39, 20136, 20136, None, None,   1,     0),
    (40, 20042, 20042, None, None,   1,     0),
]

# (label, best, time_s_or_None=TL, iters, nodes, gap_pct)
OZBAYGIN_TABLE7 = [
    ("41_v1", 3203, 1249.35,   33672,  6147,  1.88),
    ("41_v2", 2133,  854.47,    7528,   805,  2.33),
    ("42_v1", 2799,    3.00,      58,     1,  0.00),
    ("42_v2", 1946, 1005.36,    6472,   743,  4.12),
    ("43_v1", 2607,    None,  111845, 14854,  3.68),
    ("43_v2", 1966,  270.72,    1644,   135,  1.09),
    ("44_v1", 2261,   98.52,    2342,   251,  2.03),
    ("44_v2", 1610,   41.59,     222,     1,  0.00),
    ("45_v1", 3217,    1.63,      34,     1,  0.00),
    ("45_v2", 2478,    9.76,      80,     1,  0.00),
    ("46_v1", 2805,    3.81,     126,     5,  0.91),
    ("46_v2", 2469,   27.37,     302,    39,  1.11),
    ("47_v1", 3339, 3710.35,   89246, 21169,  3.38),
    ("47_v2", 1946,   68.96,     556,    37,  1.22),
    ("48_v1", 3325,    1.15,      48,     1,  0.00),
    ("48_v2", 2380,  477.83,    8384,  1493,  1.65),
    ("49_v1", 3534,  104.26,    3084,   595,  1.82),
    ("49_v2", 2492,   13.62,     207,     9,  0.55),
    ("50_v1", 2752,    8.74,     114,     1,  0.00),
    ("50_v2", 2443,  164.37,    1038,   119,  0.38),
]
OZ_TL_S_VARIANTES = 7200


def _to_float(s):
    try:
        return float(s)
    except Exception:
        return None


def load_csv(path: Path, key_field: str = "instance") -> dict:
    out = {}
    if not path.exists():
        return out
    with path.open("r", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            out[row[key_field]] = row
    return out


def build_originais() -> pd.DataFrame:
    vrps = load_csv(VRPS_ORIG)
    hexa = load_csv(HEX_SUMMARY)
    rows = []
    for k, init, oz_best, oz_gap, oz_time, oz_iters, oz_nodes in OZBAYGIN_TABLE3:
        inst_name = f"instance_{k - 1}-triangle"
        v = vrps.get(inst_name, {})
        h = hexa.get(inst_name, {})

        v_cost = _to_float(v.get("best_inc"))
        v_time = _to_float(v.get("total_time_s"))
        v_optimal = v.get("optimal") == "1"

        h_cost = _to_float(h.get("total_distance"))
        h_time = _to_float(h.get("wall_time_s"))
        h_tl = int(_to_float(h.get("time_limit_s")) or 0)
        h_n = int(_to_float(h.get("n_customers")) or 0)

        gap_h_oz = (h_cost - oz_best) / oz_best * 100.0 if (h_cost and oz_best) else None
        gap_h_v = (h_cost - v_cost) / v_cost * 100.0 if (h_cost and v_cost) else None
        gap_v_oz = (v_cost - oz_best) / oz_best * 100.0 if (v_cost and oz_best) else None

        rows.append({
            "Inst.": k,
            "n": h_n,
            "Ozb. Custo": oz_best,
            "Ozb. Tempo (s)": "TL" if oz_time is None else round(oz_time, 2),
            "Ozb. TL (s)": OZ_TL_S_BY_K[k],
            "Ozb. Gap (%)": "--" if oz_gap is None else round(oz_gap, 2),
            "VRPS Custo": int(v_cost) if v_cost else None,
            "VRPS Tempo (s)": round(v_time, 2) if v_time else None,
            "VRPS Otimo?": "sim" if v_optimal else "nao",
            "HXLY Custo": int(h_cost) if h_cost else None,
            "HXLY Tempo (s)": round(h_time, 1) if h_time else None,
            "HXLY TL (s)": h_tl,
            "Gap HXLY vs Ozb. (%)": round(gap_h_oz, 2) if gap_h_oz is not None else None,
            "Gap HXLY vs VRPS (%)": round(gap_h_v, 2) if gap_h_v is not None else None,
            "Gap VRPS vs Ozb. (%)": round(gap_v_oz, 2) if gap_v_oz is not None else None,
        })
    return pd.DataFrame(rows)


def build_variantes() -> pd.DataFrame:
    vrps_v1 = load_csv(VRPS_V1)
    vrps_v2 = load_csv(VRPS_V2)
    hexa = load_csv(HEX_SUMMARY)
    rows = []
    for label, oz_best, oz_time, oz_iters, oz_nodes, oz_gap in OZBAYGIN_TABLE7:
        m = re.match(r"(\d+)_v(\d)", label)
        k = int(m.group(1))
        var = int(m.group(2))
        if var == 1:
            inst_v = f"instance_{k - 41}"
            v = vrps_v1.get(inst_v, {})
        else:
            inst_v = f"variant2_instance_{k - 41}"
            v = vrps_v2.get(inst_v, {})
        h = hexa.get(inst_v, {})

        v_cost = _to_float(v.get("best_inc"))
        v_time = _to_float(v.get("total_time_s"))
        v_optimal = v.get("optimal") == "1"

        h_cost = _to_float(h.get("total_distance"))
        h_time = _to_float(h.get("wall_time_s"))
        h_tl = int(_to_float(h.get("time_limit_s")) or 0)
        h_n = int(_to_float(h.get("n_customers")) or 0)

        gap_h_oz = (h_cost - oz_best) / oz_best * 100.0 if (h_cost and oz_best) else None
        gap_h_v = (h_cost - v_cost) / v_cost * 100.0 if (h_cost and v_cost) else None
        gap_v_oz = (v_cost - oz_best) / oz_best * 100.0 if (v_cost and oz_best) else None

        rows.append({
            "Inst.": label,
            "n": h_n,
            "Ozb. Custo": oz_best,
            "Ozb. Tempo (s)": "TL" if oz_time is None else round(oz_time, 2),
            "Ozb. TL (s)": OZ_TL_S_VARIANTES,
            "Ozb. Gap (%)": round(oz_gap, 2),
            "VRPS Custo": int(v_cost) if v_cost else None,
            "VRPS Tempo (s)": round(v_time, 2) if v_time else None,
            "VRPS Otimo?": "sim" if v_optimal else "nao",
            "HXLY Custo": int(h_cost) if h_cost else None,
            "HXLY Tempo (s)": round(h_time, 1) if h_time else None,
            "HXLY TL (s)": h_tl,
            "Gap HXLY vs Ozb. (%)": round(gap_h_oz, 2) if gap_h_oz is not None else None,
            "Gap HXLY vs VRPS (%)": round(gap_h_v, 2) if gap_h_v is not None else None,
            "Gap VRPS vs Ozb. (%)": round(gap_v_oz, 2) if gap_v_oz is not None else None,
        })
    return pd.DataFrame(rows)


def build_resumo(df_orig: pd.DataFrame, df_var: pd.DataFrame) -> pd.DataFrame:
    def _stats(name, df, mask=None):
        d = df if mask is None else df[mask]
        gh_v = d["Gap HXLY vs VRPS (%)"].dropna()
        gh_o = d["Gap HXLY vs Ozb. (%)"].dropna()
        return {
            "Subgrupo": name,
            "# inst.": len(d),
            "VRPS bate Ozb. (#)": int(((d["VRPS Custo"] - d["Ozb. Custo"]).abs() < 0.5).sum()),
            "HXLY bate Ozb. (#)": int(((d["HXLY Custo"] - d["Ozb. Custo"]).abs() < 0.5).sum()),
            "HXLY bate VRPS (#)": int(((d["HXLY Custo"] - d["VRPS Custo"]).abs() < 0.5).sum()),
            "Gap HXLY vs VRPS medio (%)": round(gh_v.mean(), 2) if len(gh_v) else None,
            "Gap HXLY vs VRPS maximo (%)": round(gh_v.max(), 2) if len(gh_v) else None,
            "Gap HXLY vs Ozb. medio (%)": round(gh_o.mean(), 2) if len(gh_o) else None,
            "Gap HXLY vs Ozb. maximo (%)": round(gh_o.max(), 2) if len(gh_o) else None,
            "Tempo HXLY total (s)": int(pd.to_numeric(d["HXLY Tempo (s)"], errors="coerce").sum()),
            "Tempo VRPS total (s)": round(pd.to_numeric(d["VRPS Tempo (s)"], errors="coerce").sum(), 1),
        }

    rows = [
        _stats("Originais (40)", df_orig),
        _stats("Originais n=15  (1..5)",   df_orig, df_orig["n"] == 15),
        _stats("Originais n=20  (6..10)",  df_orig, df_orig["n"] == 20),
        _stats("Originais n=30  (11..20)", df_orig, df_orig["n"] == 30),
        _stats("Originais n=60  (21..30)", df_orig, df_orig["n"] == 60),
        _stats("Originais n=120 (31..40)", df_orig, df_orig["n"] == 120),
        _stats("Variantes v1 (10)", df_var, df_var["Inst."].str.endswith("_v1")),
        _stats("Variantes v2 (10)", df_var, df_var["Inst."].str.endswith("_v2")),
        _stats("Variantes (20)", df_var),
    ]
    return pd.DataFrame(rows)


def build_notas() -> pd.DataFrame:
    return pd.DataFrame({
        "Item": [
            "Mapeamento (Tabela 3)",
            "Mapeamento (Tabela 7)",
            "Ozb. Custo / Tempo / Gap",
            "Ozb. TL (s)",
            "VRPS Custo / Tempo",
            "VRPS Otimo?",
            "HXLY Custo / Tempo / TL",
            "Gap HXLY vs Ozb. (%)",
            "Gap HXLY vs VRPS (%)",
            "Gap VRPS vs Ozb. (%)",
            "'bate' (#)",
            "Hardware",
            "TL adotado no Hexaly",
            "Hexaly: motor de fundo",
            "Limitacao do Hexaly em VRPRDL",
        ],
        "Valor / Observacao": [
            "instance_{k-1}-triangle (nosso) <-> linha k da Tabela 3 do Ozbaygin (k=1..40).",
            "variant_1/instance_K (nosso) <-> {41+K}_v1; variant_2/variant2_instance_K <-> {41+K}_v2 (K=0..9).",
            "Lidos diretamente da Tabela 3 / Tabela 7 do artigo Ozbaygin et al. (2017). 'TL' indica que o solver atingiu o time-limit sem provar otimalidade.",
            "2 h (n<=60) / 6 h (n=120) no artigo.",
            "best_inc e total_time_s reportados pelo VrpSolver (Bapcod) na nossa execucao.",
            "VrpSolver retornou flag optimal=1 (Branch-Cut-and-Price exato).",
            "total_distance, wall_time_s e time_limit_s do batch do Hexaly. TL adotado: 60 s para n<=30 e variantes; 300 s para n=60; 1200 s para n=120 (originais).",
            "(HXLY - Ozb)/Ozb * 100. Positivo = nossa solucao pelo Hexaly esta acima do melhor reportado pelo artigo.",
            "(HXLY - VRPS)/VRPS * 100.",
            "(VRPS - Ozb)/Ozb * 100. Negativo = nossa execucao do VrpSolver melhorou o melhor publicado.",
            "Mesma funcao-objetivo (tolerancia 0,5).",
            "Hexaly e VrpSolver rodaram na mesma maquina do usuario (Windows + WSL). O artigo rodou em hardware diferente: a comparacao de tempo nao e pe-com-pe, mas serve de referencia de ordem de grandeza.",
            "Hexaly e anytime/heuristico para esta formulacao (model.list + lambda); ele consome todo o TL configurado, mesmo apos atingir o otimo.",
            "Hexaly 14.5 com motor de Local Search autonomo; o presolve direciona para o motor heuristico devido aos construtos model.list, model.partition e lambdas com indices de decisao.",
            "Nao gera lower bound apertado (gap permanece em 100% para custo). Por isso o selo de 'otimo' e do VrpSolver; o Hexaly entra como baseline anytime rapido.",
        ],
    })


def write_xlsx(df_orig: pd.DataFrame, df_var: pd.DataFrame,
               df_resumo: pd.DataFrame, df_notas: pd.DataFrame) -> None:
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        df_orig.to_excel(writer, sheet_name="Originais (Tab.3)", index=False)
        df_var.to_excel(writer, sheet_name="Variantes (Tab.7)", index=False)
        df_resumo.to_excel(writer, sheet_name="Resumo", index=False)
        df_notas.to_excel(writer, sheet_name="Notas", index=False)

    wb = load_workbook(OUT_XLSX)
    header_fill = PatternFill("solid", fgColor="DCE6F1")
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = bold
            cell.alignment = center
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[letter]:
                v = "" if cell.value is None else str(cell.value)
                if len(v) > max_len:
                    max_len = len(v)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 36)
        ws.freeze_panes = "B2"
    wb.save(OUT_XLSX)


def main() -> None:
    df_orig = build_originais()
    df_var = build_variantes()
    df_resumo = build_resumo(df_orig, df_var)
    df_notas = build_notas()

    write_xlsx(df_orig, df_var, df_resumo, df_notas)
    print(f"OK - arquivo gerado: {OUT_XLSX}")
    print()
    print("=== RESUMO ===")
    cols = [
        "Subgrupo", "# inst.",
        "VRPS bate Ozb. (#)", "HXLY bate Ozb. (#)", "HXLY bate VRPS (#)",
        "Gap HXLY vs VRPS medio (%)", "Gap HXLY vs VRPS maximo (%)",
        "Gap HXLY vs Ozb. medio (%)",  "Gap HXLY vs Ozb. maximo (%)",
    ]
    print(df_resumo[cols].to_string(index=False))


if __name__ == "__main__":
    main()
