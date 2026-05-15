"""
Gera tabela3_comparativo.xlsx:
- Aba 1 "VRPRDL - nosso (VRPSolver)"  : resultados desta execução + #colunas + Δtempo%
- Aba 2 "Ozbaygin 2017 - Table 3"     : valores publicados
- Aba 3 "Notas"                        : metodologia / legendas

Mapa: instance_{k-1}-triangle (nossa) <-> Table 3 row "k" (artigo).
"""

from pathlib import Path
import re
import csv

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(r"C:\Users\porin\OneDrive\Documentos\Python-Mestrado\Modelagem Matemática\Programação Inteira - Uchoa\Problema VRPRDL\VRPRDL_VRPSolver_")
LOG_DIR = ROOT / "logs"
SUMMARY_CSV = LOG_DIR / "_summary.csv"
OUT_XLSX = ROOT / "tabela3_comparativo.xlsx"


OZBAYGIN_TL_HOURS = {
    **{k: 2 for k in range(1, 31)},
    **{k: 6 for k in range(31, 41)},
}
OZBAYGIN_TL_S = {k: h * 3600 for k, h in OZBAYGIN_TL_HOURS.items()}

OZBAYGIN_TABLE3 = [
    (1,  2074, 901,   0.0,   0.58,   39,    1),
    (2,  2316, 1286,  0.0,   0.27,   29,    1),
    (3,  2234, 991,   0.0,   0.43,   30,    1),
    (4,  1982, 1062,  0.0,   0.27,   30,    1),
    (5,  3322, 1832,  0.0,   0.04,   26,    1),
    (6,  3328, 1294,  1.73,  2.30,   245,   31),
    (7,  3204, 1155,  1.29,  30.47,  773,   70),
    (8,  3170, 1455,  0.0,   0.18,   39,    1),
    (9,  2838, 1260,  1.86,  3.32,   307,   32),
    (10, 3270, 1684,  0.0,   0.55,   33,    1),
    (11, 4932, 1922,  0.31,  14.28,  283,   27),
    (12, 4610, 2324,  2.52,  120.38, 16914, 2707),
    (13, 4868, 1747,  0.0,   23.33,  101,   1),
    (14, 4084, 1273,  0.0,   30.83,  100,   1),
    (15, 4656, 1694,  0.0,   22.78,  97,    1),
    (16, 4770, 1938,  0.0,   57.13,  82,    1),
    (17, 4502, 1965,  0.10,  5.75,   117,   3),
    (18, 5392, 1827,  0.0,   2.10,   83,    1),
    (19, 5286, 2083,  2.46,  93.55,  4140,  413),
    (20, 4236, 1822,  0.0,   78.70,  103,   1),
    (21, 10374, 3761, 0.0,   606.60, 216,   1),
    (22, 9316, 2828,  0.0,   272.43, 357,   3),
    (23, 10326, 4440, 0.01,  238.27, 208,   3),
    (24, 10536, 3378, 0.0,   382.37, 250,   1),
    (25, 9784, 9784,  None,  None,   1,     0),
    (26, 10822, 4536, 0.0,   78.61,  158,   1),
    (27, 10634, 2865, 0.0,   296.23, 242,   1),
    (28, 9450, 9450,  None,  None,   1,     0),
    (29, 10988, 3964, 0.73,  None,   82536, 11593),
    (30, 11840, 4107, 0.0,   45.22,  177,   1),
    (31, 18142, 18142, None, None,   1,     0),
    (32, 19514, 19514, None, None,   1,     0),
    (33, 18008, 18008, None, None,   4,     0),
    (34, 19880, 19880, None, None,   2,     0),
    (35, 19196, 19196, None, None,   1,     0),
    (36, 21772, 21772, None, None,   3,     0),
    (37, 20010, 20010, None, None,   2,     0),
    (38, 20032, 20032, None, None,   480,   0),
    (39, 20136, 20136, None, None,   1,     0),
    (40, 20042, 20042, None, None,   1,     0),
]


def parse_log(path: Path) -> dict:
    """Extrai do log do VrpSolver:
    - iterations: nº de iterações de column generation (linhas <DWph=)
    - columns_generated: soma de <nCl=N> em todas as iterações (total de colunas
      geradas acumulado, incluindo as posteriormente removidas do pool)
    - columns_pool_peak: máximo observado em 'N columns (...)' (pico do pool,
      portanto <= columns_generated).
    """
    cg_iters = 0
    total_cols = 0
    pool_peak = 0
    rx_cg = re.compile(r"<DWph=\s*\d+>\s*<it=\s*\d+>.*?<nCl=\s*(\d+)>")
    rx_pool = re.compile(r"(\d+)\s+columns\s+\(")
    with path.open("r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            m_cg = rx_cg.search(line)
            if m_cg:
                cg_iters += 1
                total_cols += int(m_cg.group(1))
            m_pool = rx_pool.search(line)
            if m_pool:
                n = int(m_pool.group(1))
                if n > pool_peak:
                    pool_peak = n
    return {
        "iterations": cg_iters,
        "columns_generated": total_cols,
        "columns_pool_peak": pool_peak,
    }


def _to_float(s):
    try:
        return float(s)
    except Exception:
        return None


def load_our_summary():
    rows = []
    with SUMMARY_CSV.open("r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for r in reader:
            name = r["instance"]
            m = re.match(r"instance_(\d+)-triangle", name)
            if not m:
                continue
            k = int(m.group(1)) + 1
            row = {
                "instance_id": k,
                "instance_name": name,
                "root_db": _to_float(r.get("root_db")),
                "best_db": _to_float(r.get("best_db")),
                "best_inc": _to_float(r.get("best_inc")),
                "total_time_s": _to_float(r.get("total_time_s")),
                "nb_nodes": int(r.get("nb_nodes", 0) or 0),
                "optimal": r.get("optimal") == "1",
                "wall_time_s": _to_float(r.get("wall_time_s")),
            }
            log_path = LOG_DIR / f"{name}.log"
            if log_path.exists():
                row.update(parse_log(log_path))
            else:
                row.update({"iterations": None, "columns_generated": None, "columns_pool_peak": None})
            rows.append(row)
    rows.sort(key=lambda x: x["instance_id"])
    return rows


def build_dataframes():
    ours = load_our_summary()
    oz_by_id = {r[0]: r for r in OZBAYGIN_TABLE3}

    our_records = []
    for r in ours:
        k = r["instance_id"]
        oz = oz_by_id.get(k)
        oz_time = oz[4] if oz else None
        our_time = r["total_time_s"]
        if our_time is not None and oz_time is not None and oz_time > 0:
            dt_pct = (our_time - oz_time) / oz_time * 100.0
            dt_pct_str = round(dt_pct, 2)
        elif oz_time is None and oz is not None:
            dt_pct_str = f"TL no artigo ({OZBAYGIN_TL_S[k]}s)"
        else:
            dt_pct_str = None

        inc = r["best_inc"]
        root = r["root_db"]
        gap = None
        if inc is not None and root is not None and root > 0:
            gap = round((inc - root) / root * 100.0, 4)

        our_records.append(
            {
                "Instance": k,
                "Initial solution": "N/A",
                "Best solution": int(inc) if inc is not None else None,
                "Integrality gap (%)": gap,
                "Solution time (s)": round(r["total_time_s"], 2) if r["total_time_s"] is not None else None,
                "Iterations (CG)": r["iterations"],
                "Nodes": r["nb_nodes"],
                "Columns generated (total)": r["columns_generated"],
                "Columns pool (peak)": r["columns_pool_peak"],
                "Δ Time vs Özbaygın (%)": dt_pct_str,
                "Root LB": round(root, 2) if root is not None else None,
                "Optimal?": "yes" if r["optimal"] else "no",
                "Wall time (s)": int(r["wall_time_s"]) if r["wall_time_s"] is not None else None,
            }
        )

    df_ours = pd.DataFrame(our_records)

    oz_records = []
    for k, init, best, gap, time_s, iters, nodes in OZBAYGIN_TABLE3:
        oz_records.append(
            {
                "Instance": k,
                "Initial solution": init,
                "Best solution": best,
                "Integrality gap (%)": "–" if gap is None else gap,
                "Solution time (s)": "TL" if time_s is None else time_s,
                "Iterations": iters,
                "Nodes": nodes,
                "Time limit (s)": OZBAYGIN_TL_S[k],
            }
        )
    df_oz = pd.DataFrame(oz_records)

    return df_ours, df_oz


def write_xlsx(df_ours: pd.DataFrame, df_oz: pd.DataFrame):
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        df_ours.to_excel(writer, sheet_name="VRPRDL - nosso (VRPSolver)", index=False)
        df_oz.to_excel(writer, sheet_name="Ozbaygin 2017 - Table 3", index=False)

        notas = pd.DataFrame(
            {
                "Item": [
                    "Mapeamento de instâncias",
                    "Initial solution",
                    "Best solution",
                    "Integrality gap (%)",
                    "Solution time (s)",
                    "Iterations (nosso)",
                    "Iterations (artigo)",
                    "Nodes",
                    "Columns generated (total)",
                    "Columns pool (peak)",
                    "Δ Time vs Özbaygın (%)",
                    "Time limit do artigo",
                    "Time limit nosso batch",
                    "Hardware",
                    "Formulação",
                ],
                "Valor / Observação": [
                    "instance_{k-1}-triangle (nosso) ↔ linha k da Tabela 3 (artigo, 1..40)",
                    "N/A no nosso setup: rodamos sem heurística primal (UB = 1e12)",
                    "Custo incumbente final = :bcRecBestInc do VrpSolver",
                    "(best_inc - root_db)/root_db * 100; root_db = :bcRecRootDb",
                    ":bcTimeMain do VrpSolver (tempo total de solve)",
                    "Contagem de linhas '<DWph=' no log do VrpSolver (iterações de CG)",
                    "Iterations reportadas pela implementação de Özbaygın et al. 2017",
                    ":bcCountNodeProc (nós processados no B&P)",
                    "Soma de <nCl=N> em todas as iterações de CG (total acumulado de colunas geradas)",
                    "Máximo observado de 'N columns (...)' no log (pico do tamanho do pool ao longo da execução)",
                    "(nosso - artigo)/artigo * 100.  Para instâncias TL no artigo: marcado como 'TL no artigo (Xs)'",
                    "2h (instâncias 1..30, n <= 60) e 6h (instâncias 31..40, n = 120)",
                    "3h por instância (todas resolveram dentro do limite)",
                    "Nossa execução rodou no hardware do usuário (WSL). Artigo rodou em outro hardware e NÃO é comparação pé-com-pé de tempo.",
                    "VrpSolver (Pessoa/Sadykov/Uchoa) com packing sets por cliente; recursos principais capacidade e tempo.",
                ],
            }
        )
        notas.to_excel(writer, sheet_name="Notas", index=False)

    wb_path = OUT_XLSX
    from openpyxl import load_workbook

    wb = load_workbook(wb_path)
    header_fill = PatternFill("solid", fgColor="DCE6F1")
    header_font = Font(bold=True)

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        max_col = ws.max_column
        for col_idx in range(1, max_col + 1):
            letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[letter]:
                val = "" if cell.value is None else str(cell.value)
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 40)
        ws.freeze_panes = "A2"

    wb.save(wb_path)


def main():
    df_ours, df_oz = build_dataframes()
    print("Nosso (primeiras linhas):")
    print(df_ours.head())
    print()
    print("Artigo (primeiras linhas):")
    print(df_oz.head())
    print()
    write_xlsx(df_ours, df_oz)
    print(f"Arquivo gerado: {OUT_XLSX}")


if __name__ == "__main__":
    main()
